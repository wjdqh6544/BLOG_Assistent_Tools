from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import urllib.request
import time, os

FILE = "chart_data.xlsx"

def main():
    URL = input("\nEnter the URL of the article.\n> ")
    time.sleep(0.5)
    print("\n------------------------------\nTrying to get Chart Data from given article URL......", end=' ')
    time.sleep(1)
    chart = getHTML(URL)
    print("Complete!")
    time.sleep(0.5)
    print("\nParsing Received Data......", end=' ')
    write_data = getChartData(chart)
    chartCnt = len(chart)
    time.sleep(0.2)
    print("Complete!")
    time.sleep(0.5)
    print("\t- The number of Chart: " + str(chartCnt))
    for i in range(chartCnt):
        print("\t- Row of Chart #" + str(i + 1) + ": " + str(write_data[i][0][0]))
    time.sleep(0.5)
    print("\nWriting parsed data into .xlsx......")
    time.sleep(0.5)
    saveToXlsx(chartCnt, write_data)
    time.sleep(0.5)
    print("Parsing Completed!")
    print("\t- File Location: " + str(os.path.dirname(__file__)) + "\\" + FILE + "\n")
    os.system('pause')


def saveToXlsx(chartCnt, write_data):
    print("\t- Create a .xlsx File......", end=' ')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "1"
    for i in range(1, chartCnt):
        workbook.create_sheet(str(i+1))
    workbook.save(FILE)
    print("Completed!\n")
    workbook = load_workbook(FILE, read_only=False, data_only=True)
    for i in range(chartCnt):
        write_wb = workbook[str(i + 1)]
        print("\t- Writing parsed data in #" + str(i + 1) + " sheet...")
        time.sleep(0.5)
        for j in range(1, write_data[i][0][0] + 1):
            write_wb.append([write_data[i][0][j], write_data[i][1][j]])
            print("\t- Writing parsed data in #" + str(i + 1) + " sheet... (" + str(j) + "/" + str(write_data[i][0][0]) + ")")
        time.sleep(0.5)
        print("\t- Writing parsed data in #" + str(i + 1) + " sheet... Completed!\n")
    workbook.save(FILE)

def getChartData(chart):
    write_data = [[[], []] for i in range(len(chart))] # name[0]: 이름 / name[1]: 점수
    for i in range(len(chart)):
        raw_data = chart[i].find_all("div", {'class': None})
        write_data[i][0].append(int(len(raw_data)/2)) # 항목 개수
        write_data[i][1].append(int(len(raw_data)/2))
        for j in range(0, len(raw_data), 2):
            write_data[i][0].append(raw_data[j].find('span', {'class': None}).text) # 이름 저장
            write_data[i][1].append(raw_data[j + 1].text) # 점수 저장
    return write_data


def getHTML(URL): ## WCCFTech 기사 주소에 접속하여 HTML 코드 가져오기
    request = urllib.request.Request(URL)
    html = urllib.request.urlopen(request)
    source = BeautifulSoup(html.read(), 'html.parser')
    chart = source.find_all("div", "chart-data") ## 차트를 구성하는 HTML 코드만 파싱하여 리턴함.
    return chart

main()